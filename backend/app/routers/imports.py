"""Bulk import endpoints — turn a spreadsheet or photo of past records into
structured, reviewable workouts, then commit them in bulk."""
import asyncio
import base64
import io
import uuid
from datetime import date
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession

from app.dependencies import get_db
from app.models.exercise import Exercise as ExerciseModel
from app.models.exercise_set import ExerciseSet as ExerciseSetModel
from app.models.session import WorkoutSession
from app.ratelimit import enforce_llm_budget
from app.config import settings
from app.services import gemini_service

router = APIRouter()

MAX_ROWS = 220
# base64 inflates ~4/3; cap the encoded string a bit above the decoded limit so
# oversized uploads are rejected with a clean 422 before we even decode them.
_MAX_BASE64_CHARS = (settings.MAX_IMPORT_BYTES * 4) // 3 + 1024


# --------------------------------------------------------------------------- #
# Schemas
# --------------------------------------------------------------------------- #
class PreviewRequest(BaseModel):
    kind: str  # "spreadsheet" | "photo"
    content_base64: str = Field(max_length=_MAX_BASE64_CHARS)
    filename: str | None = None
    mime: str | None = None


class ImportSet(BaseModel):
    reps: int | None = None
    weight: float | None = None
    weight_unit: str = "lbs"
    duration_seconds: int | None = None


class ImportExercise(BaseModel):
    name: str
    muscle_group: str | None = None
    notes: str | None = None
    sets: list[ImportSet] = []


class ImportWorkout(BaseModel):
    person: str | None = None
    week: int | None = None
    day: int | None = None
    day_label: str | None = None
    workout_date: str | None = None
    workout_type: str | None = None
    exercises: list[ImportExercise] = []


class PreviewResponse(BaseModel):
    source_kind: str
    detected_unit: str
    people: list[str]
    needs_start_date: bool
    total_workouts: int
    total_sets: int
    summary: str
    workouts: list[ImportWorkout]


class CommitItem(BaseModel):
    device_uuid: str
    workout_date: date
    workout_type: str | None = None
    source: str = "import"
    exercises: list[ImportExercise] = []


class CommitRequest(BaseModel):
    items: list[CommitItem] = []


class CommitResponse(BaseModel):
    created: int


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _render_sheet(ws) -> str:
    """Render a worksheet's non-empty cells into compact, LLM-readable text."""
    lines: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= MAX_ROWS:
            break
        cells = ["" if c is None else str(c) for c in row]
        while cells and cells[-1] == "":
            cells.pop()
        if cells:
            lines.append(f"r{i}: " + " | ".join(cells))
    return "\n".join(lines)


def _looks_like_training(text: str) -> bool:
    upper = text.upper()
    return ("REPS" in upper and "WEIGHT" in upper) or "SETS" in upper


def _count_sets(workouts: list[ImportWorkout]) -> int:
    return sum(len(ex.sets) for w in workouts for ex in w.exercises)


# --------------------------------------------------------------------------- #
# Preview
# --------------------------------------------------------------------------- #
@router.post(
    "/import/preview",
    response_model=PreviewResponse,
    dependencies=[Depends(enforce_llm_budget)],
)
async def import_preview(body: PreviewRequest) -> PreviewResponse:
    try:
        raw = base64.b64decode(body.content_base64)
    except Exception as exc:
        raise HTTPException(status_code=422, detail="Invalid file content") from exc

    if len(raw) > settings.MAX_IMPORT_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large (max {settings.MAX_IMPORT_BYTES // (1024 * 1024)} MB).",
        )

    if body.kind == "spreadsheet":
        return await _preview_spreadsheet(raw)
    if body.kind == "photo":
        return await _preview_photo(raw, body.mime or "image/jpeg")
    raise HTTPException(status_code=422, detail="kind must be 'spreadsheet' or 'photo'")


async def _preview_spreadsheet(raw: bytes) -> PreviewResponse:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not open spreadsheet: {exc}") from exc

    # Render each training tab, then extract all tabs concurrently.
    sheets: list[tuple[str, str]] = []
    for ws in wb.worksheets:
        grid = _render_sheet(ws)
        if grid and _looks_like_training(grid):
            sheets.append(((ws.title or "Person").strip(), grid))

    # Cap fan-out: one request must not spawn an unbounded number of Gemini
    # calls. Process at most MAX_IMPORT_SHEETS training tabs.
    truncated = len(sheets) > settings.MAX_IMPORT_SHEETS
    sheets = sheets[: settings.MAX_IMPORT_SHEETS]

    results = await asyncio.gather(
        *(gemini_service.extract_spreadsheet_sheet(grid) for _, grid in sheets),
        return_exceptions=True,
    )

    workouts: list[ImportWorkout] = []
    people: list[str] = []
    detected_unit = "lbs"

    for (person, _), result in zip(sheets, results):
        if isinstance(result, Exception) or not isinstance(result, dict):
            continue
        detected_unit = result.get("unit") or detected_unit
        sheet_workouts = result.get("workouts", []) or []
        if sheet_workouts and person not in people:
            people.append(person)
        for w in sheet_workouts:
            w["person"] = person
            try:
                workouts.append(ImportWorkout(**w))
            except Exception:
                continue

    if not workouts:
        raise HTTPException(
            status_code=422,
            detail="No logged workouts found in that spreadsheet.",
        )

    total_sets = _count_sets(workouts)
    weeks = {w.week for w in workouts if w.week}
    week_note = f", {len(weeks)} weeks" if weeks else ""
    people_note = " & ".join(people) if people else "1 person"
    summary = (
        f"{people_note}{week_note}, {len(workouts)} workouts, {total_sets} sets"
    )
    if truncated:
        summary += (
            f" (only the first {settings.MAX_IMPORT_SHEETS} tabs were imported)"
        )
    return PreviewResponse(
        source_kind="spreadsheet",
        detected_unit=detected_unit,
        people=people,
        needs_start_date=True,
        total_workouts=len(workouts),
        total_sets=total_sets,
        summary=summary,
        workouts=workouts,
    )


async def _preview_photo(raw: bytes, mime: str) -> PreviewResponse:
    result = await gemini_service.extract_photo(raw, mime)
    unit = result.get("unit") or "lbs"
    workouts: list[ImportWorkout] = []
    for w in result.get("workouts", []) or []:
        try:
            workouts.append(ImportWorkout(**w))
        except Exception:
            continue
    if not workouts:
        raise HTTPException(status_code=422, detail="No workouts found in that photo.")
    total_sets = _count_sets(workouts)
    summary = f"{len(workouts)} workout(s), {total_sets} sets"
    return PreviewResponse(
        source_kind="photo",
        detected_unit=unit,
        people=[],
        needs_start_date=False,
        total_workouts=len(workouts),
        total_sets=total_sets,
        summary=summary,
        workouts=workouts,
    )


# --------------------------------------------------------------------------- #
# Commit
# --------------------------------------------------------------------------- #
@router.post("/import/commit", response_model=CommitResponse, status_code=201)
async def import_commit(
    body: CommitRequest,
    db: Annotated[AsyncSession, Depends(get_db)],
) -> CommitResponse:
    created = 0
    for item in body.items:
        try:
            device_uuid = str(uuid.UUID(str(item.device_uuid)))
        except ValueError:
            raise HTTPException(status_code=422, detail="invalid device_uuid")
        if not item.exercises:
            continue

        workout = WorkoutSession(
            device_uuid=device_uuid,
            workout_date=item.workout_date,
            source=item.source,
            workout_type=item.workout_type,
        )
        db.add(workout)
        await db.flush()

        for ex_idx, ex in enumerate(item.exercises):
            exercise = ExerciseModel(
                session_id=workout.session_id,
                exercise_order=ex_idx,
                name=ex.name,
                muscle_group=ex.muscle_group,
                notes=ex.notes,
            )
            db.add(exercise)
            await db.flush()
            for s_idx, s in enumerate(ex.sets):
                db.add(
                    ExerciseSetModel(
                        exercise_id=exercise.exercise_id,
                        set_number=s_idx + 1,
                        reps=s.reps,
                        weight=s.weight,
                        weight_unit=s.weight_unit,
                        duration_seconds=s.duration_seconds,
                    )
                )
        created += 1

    await db.flush()
    return CommitResponse(created=created)
